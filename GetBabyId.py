import os
import random
import time
import requests
from openpyxl.utils.exceptions import IllegalCharacterError
from requests.exceptions import RequestException, ReadTimeout
import re
from openpyxl import Workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
from GetUserAgent import get_user_agent
from save_fail_href import write_to_file


def get_one_page(url):
    # 代理ip
    proxy = {
        'http': '203.174.112.13:3128',
        'http': '106.122.170.176:8118',
        'http': '61.143.17.83:808',
        'http': '117.85.105.170:808',
    }

    # 获取随机的浏览器代理
    user_agent = get_user_agent()
    print('浏览器代理:\t' + user_agent)
    headers = {
        "User-Agent": user_agent
    }
    # 请求并返回网页
    try:
        response = requests.get(url, headers=headers)
        print('状态码:\t\t' + str(response.status_code))
        if response.status_code == 200:
            return response.content.decode('utf-8')
    except ReadTimeout:
        print('Timeout')
        return None
    except ConnectionError:
        print('Connection error')
        return None
    except RequestException:
        print('Error')
        return None


# 分析列表页
def parse_one_page(html):
    pattern = re.compile('<input id="(\d+)-hid" value', re.S)
    items = re.findall(pattern, html)
    return items


def get_detail_info(items, offset):
    global ws  # 全局工作表对象
    # Catagory = []
    # Id = []
    # Name = []
    # Sex = []
    # Birth = []
    # Hight = []
    # Lost_date = []
    # Home = []
    # Lost_loc = []
    # Descr = []
    # Other = []
    # Resis_time = []
    # Volunteer = []
    item_count = 0
    for item in items:
        item_count += 1
        # 将列表页的id拼接成详情页的地址
        url = 'http://www.baobeihuijia.com/view.aspx?type=1&id=' + str(item)
        print('\033[0;34;m' + '页数:\t\t' + str(offset) + '    序号: ' + str(item_count))
        # 设置随机时间延迟爬取详情页
        random_sleep_time = random.uniform(0.2, 2.000001)
        print('\033[0;34;m' + '延迟:\t\t' + str(random_sleep_time) + 's' + '    页数:' + str(offset) + '\033[0m')
        time.sleep(random_sleep_time)
        if item_count == 12:
            print('\033[0;33;m' + '中场休息...' + '\033[0m')
            time.sleep(4)
        elif item_count == 26:
            print('\033[0;33;m' + '中场休息...' + '\033[0m')
            time.sleep(3)
        # 爬取详情页
        html = get_one_page(url)

        if html is None:
            write_to_file(url)  # 将失败的连接保存在txt
            print('\033[1;31;m' + '跳过' + '\033[0m')
            print('\033[1;31;m' + '—' * 200 + '\033[0m')
            continue  # 若爬取失败则跳过，继续怕去下一个详情页

        pattern = re.compile('<li><span>.*?</span>(.*?)</li>', re.S)
        info_items = re.findall(pattern, html)  # 解析详情页的数据

        if not info_items:
            write_to_file(url)  # 将失败的连接保存在txt
            print('\033[1;31;m' + '跳过' + '\033[0m')
            print('\033[1;31;m' + '—' * 200 + '\033[0m')
            continue  # 若爬取失败则跳过，继续怕去下一个详情页

        print('\033[0;30;m' + '宝贝详情:\t\t' + str(info_items) + '\033[0m')
        print('\033[0;37;m' + '—' * 200 + '\033[0m')
        # Catagory.append(info_items[0])
        # Id.append(info_items[1].split(">")[1].split("<")[0])
        # Name.append(info_items[2])
        # Sex.append(info_items[3])
        # Birth.append(info_items[4])
        # Hight.append(info_items[5])
        # Lost_date.append(info_items[6])
        # Home.append(info_items[7])
        # Lost_loc.append(info_items[8])
        # Other.append(info_items[9])
        # Descr.append(info_items[10])
        # Resis_time.append(info_items[11])
        # Volunteer.append(info_items[12])

        # 往excel中添加数据
        try:
            ws.append([info_items[0], info_items[1].split(">")[1].split("<")[0], info_items[2].replace('\x08', ''),
                       info_items[3], info_items[4], info_items[5], info_items[6], info_items[7],
                       info_items[8].replace('\x08', ''),
                       info_items[9].replace('\x08', ''), info_items[10].replace('\x14', ''), info_items[11],
                       info_items[12], str(offset)])
        except IllegalCharacterError:
            write_to_file(url)  # 将失败的连接保存在txt
            print('\033[1;31;m' + '字符错误，跳过' + '\033[0m')
            print('\033[1;31;m' + '—' * 200 + '\033[0m')


def main(offset):
    url = 'http://www.baobeihuijia.com/list.aspx?tid=1&sex=&photo=&page=' + str(offset)
    html = get_one_page(url)  # 获取列表页html
    if html is not None:
        items = parse_one_page(html)  # 分析列表页
        if items is not None:
            get_detail_info(items, offset)
        else:
            print('\033[1;31;m' + '详情页返回空' + '\033[0m')
    else:
        print('\033[1;31;m' + '列表页返回空' + '\033[0m')


if __name__ == '__main__':
    if os.path.exists('D:\\Project\\PyCharm\\BabyGoHome\\test.xlsx'):
        print('往excel中添加数据')
        wb = load_workbook('test.xlsx')
        ws = wb.active  # 获取当前正在操作的表对象
    else:
        print('新建excel并写入数据')
        #  创建Excel表并写入数据
        wb = Workbook()  # 创建Excel对象
        ws = wb.active  # 获取当前正在操作的表对象
        # 往表中写入标题行,以列表形式写入
        ws.append(['寻亲类别', '寻亲编号', '姓名', '性别', '出生日期', '失踪时身高', '失踪时间', '失踪人所在地',
                   '失踪地点', '寻亲者特征描述', '其他资料', '注册时间', '跟进志愿者'])
    for i in range(545, 1292):
        print(
            '\033[1;32;m' + '###########################开始写入: 第' + str(i) + '页###########################' + '\033[0m')
        main(i)
        wb.save('test.xlsx')  # 存入一张列表页的中所有的失踪人详细信息后，保存为test.xlsx
        print('\033[1;32;m' + '###########################写入成功: 第' + str(
            i) + '页###########################\n' + '\033[0m')
