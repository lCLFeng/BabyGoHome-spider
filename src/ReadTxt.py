import os

from openpyxl import load_workbook, Workbook

import GetBabyId


def read_fail_href():
    if os.path.exists('D:\\Project\\PyCharm\\BabyGoHome\\test.xlsx'):
        print('往excel中添加数据')
        wb = load_workbook('test.xlsx')
        ws = wb.active  # 获取当前正在操作的表对象
    else:
        print('新建excel并写入数据')
        #  创建Excel表并写入数据
        wb = Workbook()  # 创建Excel对象
        ws = wb.active  # 获取当前正在操作的表对象
        # 往表中写入标题行,以列表形式写入
        ws.append(['寻亲类别', '寻亲编号', '姓名', '性别', '出生日期', '失踪时身高', '失踪时间', '失踪人所在地',
                   '失踪地点', '寻亲者特征描述', '其他资料', '注册时间', '跟进志愿者'])
    fail_href = []
    with open('FailHref.txt') as f:  # with open()打开文件并将其复制到变量f。文件在语句结束后被自动关闭，即是是由异常引起的结束
        while True:
            line = f.readline()
            if not line:
                break
            print(line)
            fail_href.append(line)
            if 'page' in line:
                html = GetBabyId.get_one_page(eval(line))
                items = GetBabyId.parse_one_page(html)
                GetBabyId.get_detail_info(items, 0)
                wb.save('test.xlsx')  # 存入一张列表页的中所有的失踪人详细信息后，保存为test.xlsx


if __name__ == '__main__':
    read_fail_href()
    print('success')
